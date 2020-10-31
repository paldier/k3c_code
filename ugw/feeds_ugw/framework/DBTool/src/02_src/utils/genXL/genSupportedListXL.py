###################################################################################
#        Copyright (c) 2014 
#
#        Lantiq Beteiligungs-GmbH & Co. KG 
#
#        Lilienthalstrasse 15, 85579 Neubiberg, Germany  
#
#        For licensing information, see the file 'LICENSE' in the root folder of 
#
#        this software module.
###################################################################################

'''Generates EXCEL file from XML

Reads all the input control XML files and generates a single EXCEL file by collection below
elements from all the XML files

1. Object/Param name with full hierarchy
2. Value

+

attributes specified in config file

HOW TO USE
----------
python genSupportedListXL.py *_control.xml

''' 



from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
import openpyxl.styles.colors as colors
from lxml import etree
import yaml
import sys

#-- Colors for headings
CHOCOLATE = PatternFill(start_color = 'D2691E', end_color='FFEE1111', fill_type='solid')
DARKGRAY = PatternFill(start_color = '696969', end_color='FFEE1111', fill_type='solid')

#-- Colors for names
BROWN = PatternFill(start_color = 'CD853F', end_color='FFEE1111', fill_type='solid')
TAN = PatternFill(start_color = 'D2B48C', end_color='FFEE1111', fill_type='solid')

#-- Colors for attributes and values
GREY = PatternFill(start_color = '778899', end_color='FFEE1111', fill_type='solid')
SILVER = PatternFill(start_color = 'C0C0C0', end_color='FFEE1111', fill_type='solid')

def create_workbook():
	'''
	Create a new workbook
	'''
	wb = Workbook()
	return wb

def create_worksheet(workbook, sheet_name = "UGW 7.1 SUPPORTED OBJECTS"):
	'''
	Create a worksheet with a name

	workbook(IN) -> Workbook in which you want to create the sheet in
	sheet_name(IN) -> Sheet name
	'''

	ws = workbook.active
	ws.title = sheet_name
	return ws

def get_full_path(node):
	'''
	Function for fully qualifying a node name to generate its full hierarchy

	node(IN) -> Node for which you want to return full hierarchy string
    '''
	sDefine = '' + node.tag
	while node.getparent() != None:
		sDefine = node.getparent().tag + '.' + sDefine
		node = node.getparent()
	return sDefine

def fill_headings_and_style(worksheet, attribute_dict):
	'''
	Fill the heading names and style

	worksheet(IN) 		 -> Worksheet in which you want to fill headings
	attribute_dict(IN) -> Mapping with column name and heading name in
							that column
	'''

	name_value_dict = {'A':'name', 'C':'value'}	
	for column in name_value_dict:
		heading = column + str(3)
		worksheet[heading].font = Font(size = 16, bold = True)
		worksheet[heading] = name_value_dict[column].upper()
		worksheet[heading].alignment = Alignment(horizontal='center')

	for column in attribute_dict:
		heading = column + str(3)
		worksheet[heading].font = Font(size = 16, bold = True)
		worksheet[heading] = attribute_dict[column].upper()
		worksheet[heading].alignment = Alignment(horizontal='center')

	#-- Final column = [A, B, C] + number of attributes
	final_column = chr(ord('C')+ len(attribute_dict))
	heading_size = 'A1:' + final_column + '4'
	#print('@@@@@@@@ ' + heading_size)
	
	#-- Style heading
	for row in worksheet.iter_rows(heading_size):
		for cell in row:
			if cell.column == 'A' or cell.column == 'B':
				cell.fill = CHOCOLATE
			else:
				cell.fill = DARKGRAY

def generate_excel(xml_tree, attribute_dict):
	'''
	Generates excel by taking xml as input. Generated excel will
	list following from the xml

        1. Node name with full hierarchy
        2. Node value if present
		 
		+

		attributes listed in config file
		
	xml_tree(IN) -> tree of xml from which you need to generate excel
	'''	
	wb = create_workbook()
	ws = create_worksheet(wb)

	root = xml_tree.getroot()

	i = 4
	max_column_len = 0

	#-- Create a dictonaary to find out max string length and increase column width
	column_width_dict = {'A': 0, 'C': 0}
	for key in attribute_dict:
		column_width_dict[key] = 0
	#print '@@@@@ ', column_width_dict

	fill_headings_and_style(ws, attribute_dict)

	for node in root.iter():

		if node.tag == 'Device':
			continue

		i = i + 1

		#-- Fill Object/parameter name
		name = get_full_path(node)
		ws.cell(row = i, column = 1).value = name
		
		#-- If parameter
		if len(node) == 0:
			ws.cell(row = i, column = 2).value = 'P'
			ws.cell(row = i, column = 2).alignment = Alignment(horizontal='center')
		#-- If object/service
		else:
			ws.cell(row = i, column = 1).font = Font(bold = True)
			if name.count('.') == 1:
				ws.cell(row = i, column = 2).value = 'S'
				ws.cell(row = i, column = 2).alignment = Alignment(horizontal='center')
			else:
				ws.cell(row = i, column = 2).value = 'O'
				ws.cell(row = i, column = 2).alignment = Alignment(horizontal='center')

		#-- Style your cell
		ws.cell(row = i, column = 1).alignment = Alignment(horizontal='left')

		if i % 2 == 0:
			ws.cell(row = i , column = 1).fill = BROWN
			ws.cell(row = i , column = 2).fill = BROWN
		else:
			ws.cell(row = i , column = 1).fill = TAN
			ws.cell(row = i , column = 2).fill = TAN
		
		#-- Store the length of biggest object to set column width 
		if len(name) > column_width_dict['A']:
			column_width_dict['A'] = len(name)

		#-- Fill 'text' value
		ws.cell(row = i, column = 3).value = node.text
		
		ws.cell(row = i, column = 3).alignment = Alignment(horizontal='center')
		if i % 2 == 0:
			ws.cell(row = i , column = 3).fill = GREY
		else:
			ws.cell(row = i , column = 3).fill = SILVER

		if len(node.tag) > column_width_dict['C']:
			column_width_dict['C'] = len(node.tag)

		#-- Fill attributes	
		for key in attribute_dict:
			ws.cell(key + str(i)).value = node.get(attribute_dict[key])
			ws.cell(key + str(i)).alignment = Alignment(horizontal='center')

			if i % 2 == 0:
				ws.cell(key + str(i)).fill = GREY
			else:
				ws.cell(key + str(i)).fill = SILVER	

			if len(node.tag) > column_width_dict[key]:
				column_width_dict[key] = len(node.tag)	

	#-- Increase column width to that of the biggest name in the column
	for column in column_width_dict:
		ws.column_dimensions[column].width = column_width_dict[column]
	return wb

def stitch_xmls(xmlList):
	'''
	Stitches all the xmls in xmlList into a single xml file and
	returns the tree

	xmlList(IN) -> List of XMLs
	'''
	parser = etree.XMLParser(remove_blank_text = True)

	count = 0
	tree = None
	for filename in xmlList:
		#filename = destPath + '/' + filename.split('/')[-1]
		if count == 0:
			tree = etree.parse(filename)
			count = 1
			continue
		elif count != 0 and tree != None:
			tree_next = etree.parse(filename, parser)
			root = tree_next.getroot()
			tree.getroot().append(tree_next.getroot())
	etree.strip_tags(tree.getroot(), 'Device')
	return tree

if __name__ == '__main__':
	attribute_dict = {}
	with open('./config.yaml', 'r') as f:
		doc = yaml.load(f)
		index = 'C'
		for attrib in doc['attributes']:
			index = chr(ord(index)+1)
			attribute_dict[index] = attrib
	print attribute_dict	
	xml_tree = stitch_xmls(sys.argv[1:])
	workbook = generate_excel(xml_tree, attribute_dict)
	workbook.save('./ugw_supported.xlsx')
