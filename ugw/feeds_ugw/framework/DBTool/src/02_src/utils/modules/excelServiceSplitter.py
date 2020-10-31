
#********************************************************************************
#
#  Copyright (c) 2015
#  Lantiq Beteiligungs-GmbH & Co. KG
#  Lilienthalstrasse 15, 85579 Neubiberg, Germany 
#  For licensing information, see the file 'LICENSE' in the root folder of
#  this software module.
#
#********************************************************************************/

from openpyxl import load_workbook
from openpyxl import Workbook
import sys

sys.path.append('../common')
from common import *


def excelServiceSplitter(inputFile):
	wbLarge = load_workbook(filename = inputFile)
	sheetsTotal = wbLarge.get_sheet_names()

	for sheet in sheetsTotal:
		ROW=0
		wsLarge = wbLarge.get_sheet_by_name(sheet)
		
		wbService = Workbook()
		wsService = wbService.active
		wsService.title=sheet
		
		for row in wsLarge.iter_rows():
			COL=1
			ROW+=1
			for cell in row:
				wsService.cell(row=ROW,column=COL).value = cell.value
				COL+=1
		
		file=excel_services_path+sheet+".xlsx"
		print "Created: ",file
		wbService.save(file)
	
	return 0