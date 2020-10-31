
#********************************************************************************
#
#  Copyright (c) 2015
#  Lantiq Beteiligungs-GmbH & Co. KG
#  Lilienthalstrasse 15, 85579 Neubiberg, Germany 
#  For licensing information, see the file 'LICENSE' in the root folder of
#  this software module.
#
#********************************************************************************/

from xml.dom.minidom import parse
import xml.dom.minidom
import sys,os
from xml.dom.minidom import Document

from pathCorrect import pathCorrect

from openpyxl import Workbook,load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Style, Font
from openpyxl.writer.dump_worksheet import WriteOnlyCell
from openpyxl.comments import Comment
import getAttributesList
import XmlExlWatermark

from common import *
from excel_reference_values import *

class instanceCreationDataEXL():
	
	def __init__(self):
		pass

	def commentMultiInst(self,file):
		wb=load_workbook(file)
		allSheets = wb.get_sheet_names()
		#print sheets
		
		for sheet in allSheets:
			ws = wb.get_sheet_by_name(name = sheet)
			
			index=0
			while (ws.cell(row=PATH_ROW+index,column=PATH_COLUMN)):
				if ws.cell(row=PATH_ROW+index,column=PATH_COLUMN).value != None:
					if ws.cell(row=PATH_ROW+index,column=INSTANCE_COLUMN).value != '--':
						if ws.cell(row=PATH_ROW+index,column=INSTANCE_COLUMN).value != None:
							commentVal = ws.cell(row=PATH_ROW+index,column=INSTANCE_COLUMN).value
							commentWrite = Comment("MultiInst=Yes", "Author")
							ws.cell(row=PATH_ROW+index,column=PATH_COLUMN).comment = commentWrite
						
					index+=1
				else:
					break
			
			## delete INSTANCE_COLUMN
			index=0
			while (ws.cell(row=PATH_ROW+index,column=INSTANCE_COLUMN)):
				if ws.cell(row=PATH_ROW+index,column=INSTANCE_COLUMN).value != None:
					ws.cell(row=PATH_ROW+index,column=INSTANCE_COLUMN).value = None
					index+=1
				else:
					break
			ws["C2"] = None					## Instance attribute name is nullified
		
		wb.save(file)
		return 0
		
	def printLineage(self,node,rootName,lineage,mainList,attribElemList,flag='NULL'):
		#print "node=",node
		#print "rootName=",rootName
		#print "mainList(printLineage)",mainList

		if node.tagName != rootName:							
			#print node.tagName
			try:
				if node.attributes["Instance"]:
					attribVal=node.attributes["Instance"]
					newVal = node.tagName + ".{" + attribVal.value + "}"
					lineage.append(newVal)
			except KeyError:
				lineage.append(node.tagName)					
			
			return self.printLineage(node.parentNode,rootName,lineage,mainList,attribElemList)

		else:
			lineage.append(rootName)
			#print "lineage=",lineage
			lineage.reverse()
			writeString= '.'.join(lineage)			## content here to write
			mainList.append(writeString)
			#print "writeString-",writeString
			#print "mainList(inline)=",mainList
			return mainList
	
	
	def getChildNodes(self,node):
		#print "getChildNodes(node)",node
		childList = []
		for childs in node.childNodes:
			if childs.nodeType == xml.dom.Node.ELEMENT_NODE:
				childList.append(childs)
		#print "getChildNodes(childList)",childList
		return childList


	def depthFirst(self,node,rootName,mainList,attribElemList):
		root = node
		child = self.getChildNodes(root)
		#print "root(DF)=",root
		#print "child(DF)=",child
		#print "mainList(DF)",mainList
		#print "rootName(DF)=",rootName
				
		if child != []:
			#print "It's not a leaf Node"
			#randItem = random.choice(child)							## randomly pick an element from a list
			randItem = child[-1]									## pick the last child (to maintain the original order)
			return self.depthFirst(randItem,rootName,mainList,attribElemList)

		elif child == []:
			#print "\n"
			#print "It's a Leaf Node"
			lineage=[]
			attribElemList.append(node)
			#print "node(leaf)=",node
			mainList=self.printLineage(root,rootName,lineage,mainList,attribElemList,"firstAttrib")
			#print "mainList(received)",mainList
			#print "attribElemList(received)",attribElemList

			if root.tagName == rootName:
				return mainList
			
			node = root.parentNode									## get Parent Node
			root.parentNode.removeChild(root)						## remove current node
			#print "Parent node=",node
			return self.depthFirst(node,rootName,mainList,attribElemList)

	def main(self,file):
		print "Processing: ",file
		print "####: ", os.path.basename(file)
		print "####: ",os.path.dirname(file)
		
		file_name = os.path.basename(file)
		baseDir = os.path.dirname(file)
		
		DOMTree = xml.dom.minidom.parse(file)
		
		
		''' Get the root node ('Device') '''
		for child in DOMTree.childNodes:
			if child.nodeType == xml.dom.Node.ELEMENT_NODE:
				device = child
		
		
		''' Get Comment Node '''
		for child in DOMTree.childNodes:
			if child.nodeType == xml.dom.Node.COMMENT_NODE:
				comment = child


		try:
			watermarkText = comment.data
		except UnboundLocalError:
			print "No watermark exists in this file"
			watermarkText = "No watermark"

		
		wb = Workbook()
		ws = wb.active
		ws.title=device.tagName
		
		''' create excel sheets in a workbook '''
		for services in device.childNodes:
			if services.nodeType == xml.dom.Node.ELEMENT_NODE:
				ws = wb.create_sheet()
				ws.title = services.tagName

		''' get all the attribute names available in the XML '''
		attribNameList=getAttributesList.main(file)
		
		try:
			attribNameList.remove('Instance')
		except ValueError:
			print "No instance exists"
			pass
		
		print attribNameList
		
		#####################################################################################
		''' ###--- DEVICE :: First sheet of excel, root (or Device in this XML file) '''
		ws = wb.get_sheet_by_name(name = device.tagName)
		ws['A1'] = 'PATH'
		
		offsetRoot=0
		for attribNames in attribNameList:
			ws.cell(row=ATTRIB_NAME_ROW_DATA,column=ATTRIB_NAME_COLUMN_DATA+offsetRoot).value = attribNames
			offsetRoot+=1
		
		XmlExlWatermark.insertInExcel(ws,PATH_ROW,PATH_COLUMN,watermarkText)
		ws.cell(row=PATH_ROW,column=PATH_COLUMN).value = device.tagName
		attributes=device.attributes.items()
		ws.cell(row=ATTRIB_NAME_ROW_DATA,column=2).value='defaultValue'
		ws.cell(row=PATH_ROW,column=2).value='--'
		
		for colDumm in range(len(attribNameList)):
			ws.cell(row=PATH_ROW,column=colDumm+ATTRIB_NAME_COLUMN_DATA).value='--'

		for items in attributes:
			nameColIndex=0
			name=items[0]
			value=items[1]
			
			while ws.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_NAME_COLUMN_DATA+nameColIndex).value != None:
				nameColIndex+=1
			
			ws.cell(row=PATH_ROW,column=ATTRIB_VALUE_COLUMN_DATA+nameColIndex).value=value
		
		###----------------------------------------------------------------------------------###
		########################################################################################
		
		''' start populating data in each excel sheet of a workbook. '''
		for services in device.childNodes:
			if services.nodeType == xml.dom.Node.ELEMENT_NODE:
				ws = wb.get_sheet_by_name(name = services.tagName)
				ws['A1'] = 'PATH'
				offset=0							## write attribute Names

				for attribNames in attribNameList:
					ws.cell(row=ATTRIB_NAME_ROW_DATA,column=ATTRIB_NAME_COLUMN_DATA+offset).value = attribNames
					offset+=1

				ws.cell(row=ATTRIB_NAME_ROW_DATA,column=2).value = 'defaultValue'
				
				root=services
				rootName=services.tagName
				mainList=[]
				attribElemList=[]
				listToWriteSheet = self.depthFirst(root,rootName,mainList,attribElemList)
				#print "listToWriteSheet=",listToWriteSheet,"\n"
				listToWriteSheet.reverse()								## two consecutive lists to be maintained in order
				attribElemList.reverse()								## attribElement List in order with listToWriteSheet
				index=0
				
				for i in listToWriteSheet:
					ws.cell(row=PATH_ROW+index,column=PATH_COLUMN).value=i							## write the element lists
					
					## write dummy values to all the attributes -- in excel sheet
					for colDumm in range(len(attribNameList)+1):
						ws.cell(row=PATH_ROW+index,column=colDumm+ATTRIB_NAME_COLUMN_DATA-1).value='--'
		
					## write the parameter Values
					try:
						elemNodeValue=attribElemList[index].firstChild.data
						if elemNodeValue != '':
							ws.cell(row=PATH_ROW+index,column=2).value=elemNodeValue
					except AttributeError:
						NodeValueDontExists=True
						
					## write the attributes
					attrList = attribElemList[index].attributes.items()
					#print "attrList=",attrList
					
					for items in attrList:
						nameColIndex=0
						name=items[0]
						value=items[1]
						
						if name != 'Instance':
						
							while ws.cell(row=ATTRIB_NAME_ROW_DATA,column=ATTRIB_NAME_COLUMN_DATA+nameColIndex).value != name:
								nameColIndex+=1
								
							#print "nameColIndex=",nameColIndex
							ws.cell(row=PATH_ROW+index,column=ATTRIB_VALUE_COLUMN_DATA+nameColIndex).value=value
						
					index+=1
		
		file = os.path.join(baseDir, file_name.split(".")[0]+".xlsx")
		print "##(toDataEXL): ",file
		wb.save(file)
		
		returnVal = self.commentMultiInst(file)
		
		if returnVal == 0:
			print "Status: Successfully created data XML: ",file,"\n"
		else:
			sys.exit(0)