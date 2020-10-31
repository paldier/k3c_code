
#********************************************************************************
#
#  Copyright (c) 2015
#  Lantiq Beteiligungs-GmbH & Co. KG
#  Lilienthalstrasse 15, 85579 Neubiberg, Germany 
#  For licensing information, see the file 'LICENSE' in the root folder of
#  this software module.
#
#********************************************************************************/

from openpyxl import load_workbook
import os,sys
from itertools import islice
from xml.etree.ElementTree import Element, SubElement, ElementTree, dump,tostring
from xml.etree import ElementTree
import re

from xml.dom.minidom import *
import xml.dom.minidom

from excel_reference_values import *
from common import *

class instanceCreation():


	def __init__(self):
		pass
	
	def writeToFile(self,file):
		f=open(file,'w')
		f.write(self.doc.toprettyxml())
		f.close()


	def addAttributes(self,node,dict):
		#print "node=",node
		for i in dict:
			if ((dict[i] != '--')):				## since 'parameterValue' is handled separately
				value=dict[i]					## value
				name=i							## name
				#print "Attrib::",name,"--",value
				
				node.setAttribute(name,str(value))					## BUG FIXED - if value is a number, it was throwing an error at setAttribute.
																	## AttributeError resolved by converting any contents to string format.


	def findInstances(self,string):
		stringList = string.split('.')
		match = re.findall(r'.{\d+}',string)

		try:
			if match != []:
				for i in match:
					#print i
					match = re.search(r'\d',i)
					new = match.group()
					newString = "_"+new
					string = string.replace(i,newString)

		except AttributeError:
			pass

		return string


	def manageInstance(self,parent,childName,instance):
		#print "Instance object"
		flag = False
		
		for childNodes in parent.childNodes:
			if childNodes.nodeType == xml.dom.Node.ELEMENT_NODE:
				if childNodes.tagName == childName:
					#print "Element exists, check for instance"
					getInstanceValue = childNodes.getAttributeNode('Instance')
					#print "getInstanceValue=",getInstanceValue.nodeValue
					if getInstanceValue.nodeValue == instance:
						flag = True
						return childNodes
		if flag == False:
			#print "Element doesn't exists\nCreate element and instance"
			childNode = self.doc.createElement(childName)
			childNode.setAttribute("Instance",str(instance))
			parent.appendChild(childNode)
			return childNode


	def checkIfElementExists(self,parent,childName):
		flag = False
		elementList = childName.split('_')
		
		if len(elementList) == 2:
			childName = elementList[0]
			instance = elementList[1]
			return self.manageInstance(parent,childName,instance)
			
		else:
			childName = elementList[0]
		
		for childNodes in parent.childNodes:
			if childNodes.nodeType == xml.dom.Node.ELEMENT_NODE:
				if childNodes.tagName == childName:
					#print "Element exists"
					flag = True
					return childNodes
		
		if flag == False:
			#print "Element doesn't exists\nCreate element"
			childNode = self.doc.createElement(childName)
			parent.appendChild(childNode)
			return childNode


	def buildXML(self,list,attrib='NULL',parent='NULL',index=0):
		depth = len(list)
		
		if depth == index:
			self.addAttributes(parent,attrib)
			return
		
		#print "\n","list=",list
		#print "attrib=",attrib
		#print "parent=",parent
		#print "index=",index
		
		element = list[index]
		parent = self.checkIfElementExists(parent,element)
		index+=1
		self.buildXML(list,attrib,parent,index)


	def main(self,excelFile):
		print "Received: ",excelFile
		print "####: ", os.path.basename(excelFile)
		print "####: ",os.path.dirname(excelFile)
		
		XML_NAME = os.path.basename(excelFile)
		baseDir = os.path.dirname(excelFile)
		wb=load_workbook(excelFile)
		self.doc = Document()
		
		
		##########################################################################################
		############ handle the root of the XML file here (in this instance, the root is 'Device')
		rootSheet=wb.get_sheet_names()[0]
		self.root=self.doc.createElement(rootSheet)							## root element to build a tree
		self.doc.appendChild(self.root)										## first child 'device' is added

		wsR=wb.get_sheet_by_name(name = rootSheet)
		attribDict={}
		
		waterMarkText = (wsR['A4'].comment).text
		print "waterMarkText: ",waterMarkText

		index=0
		while wsR.cell(row=ATTRIB_NAME_ROW_DATA,column=ATTRIB_NAME_COLUMN_DATA+index):
			if wsR.cell(row=ATTRIB_NAME_ROW_DATA,column=ATTRIB_NAME_COLUMN_DATA+index).value != None:
				attribDict[wsR.cell(row=ATTRIB_NAME_ROW_DATA,column=ATTRIB_NAME_COLUMN_DATA+index).value] = wsR.cell(row=ATTRIB_VALUE_ROW_DATA,column=ATTRIB_VALUE_COLUMN_DATA+index).value
				index+=1						## column increment for every change in attribute
			else:
				break
		
		self.addAttributes(self.root,attribDict)
		############----------------------------------------------------------------###############
		###########################################################################################
		
		workSheets= wb.get_sheet_names()
		
		for sheets in workSheets[1:]:										## Ignoring the first sheet,since its a root node
			print "service tab: ",sheets
			ws=wb.get_sheet_by_name(name = sheets)
			self.services=self.doc.createElement(sheets)
			self.root.appendChild(self.services)


			index=0
			childOfServiceList=[]
			while ws.cell(row=PATH_ROW+index,column=PATH_COLUMN):
				if ws.cell(row=PATH_ROW+index,column=PATH_COLUMN).value != None:
					childOfServiceList.append(ws.cell(row=PATH_ROW+index,column=PATH_COLUMN).value)
					index+=1
				else:
					break
			
			rowIndex=0
			
			for i in childOfServiceList:
				##this part takes the attributes and keep them in dict format
				attribDict={}
				
				index=0
				
				while (1): #ws.cell(row=ATTRIB_NAME_ROW_DATA,column=ATTRIB_NAME_COLUMN_DATA+rowIndex):
					if ws.cell(row=ATTRIB_NAME_ROW_DATA,column=ATTRIB_NAME_COLUMN_DATA+index).value != None:
						attribDict[ws.cell(row=ATTRIB_NAME_ROW_DATA,column=ATTRIB_NAME_COLUMN_DATA+index).value] = ws.cell(row=ATTRIB_VALUE_ROW_DATA+rowIndex,column=ATTRIB_VALUE_COLUMN_DATA+index).value
						index+=1						## column increment for every change in attribute
					else:
						break

				##these two lines specifically reads the 'parameterValue' column of excel sheet
				attribDict[ws.cell(row=PARAMETER_NAME_ROW_DATA,column=PARAMETER_NAME_COLUMN_DATA).value] = ws.cell(row=PARAMETER_VALUE_ROW_DATA+rowIndex,column=PARAMETER_VALUE_COLUMN_DATA).value

				i=self.findInstances(i)
				
				lineageList = i.split('.')
				self.buildXML(lineageList,attribDict,self.root)
				rowIndex+=1

		if waterMarkText != 'No watermark':
			fileType = waterMarkText.split(',')[0]
			access = waterMarkText.split(',')[1]
		else:
			fileType = 'instance'
			access = 'ReadOnly'
			
		xmlFile = os.path.join(baseDir, XML_NAME.split(".xlsx")[0] + ".xml")
		self.writeToFile(xmlFile)
		return xmlFile,fileType,access