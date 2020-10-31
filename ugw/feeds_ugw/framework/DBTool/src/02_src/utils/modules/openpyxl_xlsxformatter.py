
#********************************************************************************
#
#  Copyright (c) 2015
#  Lantiq Beteiligungs-GmbH & Co. KG
#  Lilienthalstrasse 15, 85579 Neubiberg, Germany 
#  For licensing information, see the file 'LICENSE' in the root folder of
#  this software module.
#
#********************************************************************************/

from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl.styles import Style,PatternFill,Alignment
from openpyxl.styles import Color, Fill
from openpyxl.cell import get_column_letter
import sys,os

sys.path.append('modules')
from excel_reference_values import *
from color_formatter import *

## COLOR FORMATTING
HEADING_FORMAT = Style(font=Font(color=Color(WHITE), bold=True),fill=PatternFill(patternType='solid', fgColor=Color(HEADING_FORMAT_COLOR)),alignment=Alignment(wrap_text=True, horizontal='center'))

PATH_FORMAT_0 = Style(font=Font(color=Color(BLACK), bold=True),fill=PatternFill(patternType='solid', fgColor=Color(PATH_FORMAT_0_COLOR)),alignment=Alignment(wrap_text=True))

PATH_FORMAT_1 = Style(font=Font(color=Color(BLACK), bold=True),fill=PatternFill(patternType='solid', fgColor=Color(PATH_FORMAT_1_COLOR)),alignment=Alignment(wrap_text=True))

DESCRIPTION_FORMAT_0 = Style(font=Font(color=Color('FF333333'), bold=False),fill=PatternFill(patternType='solid', fgColor=Color(DESCRIPTION_FORMAT_0_COLOR)),alignment=Alignment(wrap_text=True))

DESCRIPTION_FORMAT_1 = Style(font=Font(color=Color('FF333333'), bold=False),fill=PatternFill(patternType='solid', fgColor=Color(DESCRIPTION_FORMAT_1_COLOR)),alignment=Alignment(wrap_text=True))

ATTRIB_FORMAT_0 = Style(font=Font(color=Color('FF2B5464'), bold=False),fill=PatternFill(patternType='solid', fgColor=Color(ATTRIB_FORMAT_0_COLOR)),alignment=Alignment(wrap_text=False, horizontal='center'))

ATTRIB_FORMAT_1 = Style(font=Font(color=Color('FF663399'), bold=False),fill=PatternFill(patternType='solid', fgColor=Color(ATTRIB_FORMAT_1_COLOR)),alignment=Alignment(wrap_text=False, horizontal='center'))

def openpyxl_xlsxformatter(file):
	print "Received file(FORMATTER): ",file
	
	wb = load_workbook(file)

	for ws in wb:
		
		ws.column_dimensions["A"].width = 60.0
		try:
			ws.column_dimensions["B"].width = 60.0
		except KeyError:
			pass

		## FILLERS_PATH
		ws.cell(row=ATTRIB_NAME_ROW,column=1).style = HEADING_FORMAT
		ws.cell(row=ATTRIB_NAME_ROW+1,column=1).style = HEADING_FORMAT
		ws.cell(row=ATTRIB_NAME_ROW+2,column=1).style = HEADING_FORMAT
		
		## FILLERS_DESCRIPTION
		ws.cell(row=ATTRIB_NAME_ROW,column=2).style = HEADING_FORMAT
		ws.cell(row=ATTRIB_NAME_ROW+1,column=2).style = HEADING_FORMAT
		ws.cell(row=ATTRIB_NAME_ROW+2,column=2).style = HEADING_FORMAT

		## FORMAT THE ATTRIBUTE NAMES
		index=0-3
		while ws.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_NAME_COLUMN+index):
			if ws.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_NAME_COLUMN+index).value != None:
				ws.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_NAME_COLUMN+index).style = HEADING_FORMAT
				ws.cell(row=ATTRIB_NAME_ROW+1,column=ATTRIB_NAME_COLUMN+index).style = HEADING_FORMAT		## FILLERS
				ws.cell(row=ATTRIB_NAME_ROW+2,column=ATTRIB_NAME_COLUMN+index).style = HEADING_FORMAT		## FILLERS
				
				column_letter = get_column_letter((ATTRIB_NAME_COLUMN+index))
				ws.column_dimensions[column_letter].width = len(ws.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_NAME_COLUMN+index).value)+6
				
				index+=1
			else:
				break
		
		index=0
		FLAG=0
		while ws.cell(row=PATH_ROW+index,column=PATH_COLUMN):
			if ws.cell(row=PATH_ROW+index,column=PATH_COLUMN).value != None:
				
				if FLAG==0:
					ws.cell(row=PATH_ROW+index,column=PATH_COLUMN).style = PATH_FORMAT_0
					ws.cell(row=PATH_ROW+index,column=2).style = DESCRIPTION_FORMAT_0
					
					colIndex=-3
					while ws.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_VALUE_COLUMN+colIndex):
						if ws.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_VALUE_COLUMN+colIndex).value != None:
							ws.cell(row=PATH_ROW+index,column=ATTRIB_VALUE_COLUMN+colIndex).style = ATTRIB_FORMAT_0
							colIndex+=1
							
						else:
							break
					
					FLAG=1
				
				elif FLAG==1:
					ws.cell(row=PATH_ROW+index,column=PATH_COLUMN).style = PATH_FORMAT_1
					ws.cell(row=PATH_ROW+index,column=2).style = DESCRIPTION_FORMAT_1
					
					colIndex=-3
					while ws.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_VALUE_COLUMN+colIndex):
						if ws.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_VALUE_COLUMN+colIndex).value != None:
							ws.cell(row=PATH_ROW+index,column=ATTRIB_VALUE_COLUMN+colIndex).style = ATTRIB_FORMAT_1
							colIndex+=1
							
						else:
							break
					
					FLAG=0
					
				index+=1
			else:
				break

	wb.save(file)