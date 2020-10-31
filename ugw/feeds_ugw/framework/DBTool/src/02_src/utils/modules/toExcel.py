
#********************************************************************************
#
#  Copyright (c) 2015
#  Lantiq Beteiligungs-GmbH & Co. KG
#  Lilienthalstrasse 15, 85579 Neubiberg, Germany 
#  For licensing information, see the file 'LICENSE' in the root folder of
#  this software module.
#
#********************************************************************************/

'''
This script creates a new excel file from a fully generated XML file (fully generated XML file is the outcome of all transformations)
xlsxwriter is a module that is being used to write to excel file. 
Input to this file is an xml file

A separate worksheet is created for all the services in a workbook.
Each worksheet is dumped with the child nodes and corresponding attribute values (time being dummy values are filled in attributes)
'''

## Import python modules
from xml.dom.minidom import parse
import xml.dom.minidom
import sys,os
from xml.dom.minidom import Document

from openpyxl import Workbook,load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Style, Font
from openpyxl.writer.dump_worksheet import WriteOnlyCell
from openpyxl.comments import Comment
import getAttributesList
import XmlExlWatermark

sys.path.append('../common')
from common import *

sys.path.append('modules')
from excel_reference_values import *

from excelServiceSplitter import excelServiceSplitter
from openpyxl_xlsxformatter import openpyxl_xlsxformatter

sys.path.append(r'../action/')
from pathCorrect import pathCorrect

class toExcel():
	
	def __init__(self):
		print "Initialising: toExcel"
		sys.setrecursionlimit(10000)

	def commentMultiInst(self,file):
		wb=load_workbook(file)
		allSheets = wb.get_sheet_names()
		#print sheets
		
		for sheet in allSheets:
			ws = wb.get_sheet_by_name(name = sheet)
			
			index=0
			while (ws.cell(row=PATH_ROW+index,column=PATH_COLUMN)):
				if ws.cell(row=PATH_ROW+index,column=PATH_COLUMN).value != None:
					if ws.cell(row=PATH_ROW+index,column=INSTANCE_COLUMN).value != '--':
						if ws.cell(row=PATH_ROW+index,column=INSTANCE_COLUMN).value != None:
							commentVal = ws.cell(row=PATH_ROW+index,column=INSTANCE_COLUMN).value
							commentWrite = Comment("MultiInst=Yes", "Author")
							ws.cell(row=PATH_ROW+index,column=PATH_COLUMN).comment = commentWrite
						
					index+=1
				else:
					break
			
			''' delete INSTANCE_COLUMN '''
			index=0
			while (ws.cell(row=PATH_ROW+index,column=INSTANCE_COLUMN)):
				if ws.cell(row=PATH_ROW+index,column=INSTANCE_COLUMN).value != None:
					ws.cell(row=PATH_ROW+index,column=INSTANCE_COLUMN).value = None
					index+=1
				else:
					break
			ws["C2"] = None					## Instance attribute name is nullified
		
		wb.save(file)
		return 0

	def printLineage(self,node,rootName,lineage,mainList,attribElemList,flag='NULL'):
		#print "node=",node
		#print "rootName=",rootName
		#print "mainList(printLineage)",mainList

		if node.tagName != rootName:							
			#print node.tagName
			lineage.append(node.tagName)					
			return self.printLineage(node.parentNode,rootName,lineage,mainList,attribElemList)

		else:
			lineage.append(rootName)
			#print "lineage=",lineage
			lineage.reverse()
			writeString= '.'.join(lineage)			## content here to write
			mainList.append(writeString)
			#print "writeString-",writeString
			#print "mainList(inline)=",mainList
			return mainList

	def getChildNodes(self,node):
		#print "getChildNodes(node)",node
		childList = []
		for childs in node.childNodes:
			if childs.nodeType == xml.dom.Node.ELEMENT_NODE:
				childList.append(childs)
		#print "getChildNodes(childList)",childList
		return childList


	def depthFirst(self,node,rootName,mainList,attribElemList):
		root = node
		child = self.getChildNodes(root)
		#print "root(DF)=",root
		#print "child(DF)=",child
		#print "mainList(DF)",mainList
		#print "rootName(DF)=",rootName
				
		if child != []:
			#print "It's not a leaf Node"
			#randItem = random.choice(child)							## randomly pick an element from a list
			randItem = child[-1]									## pick the last child (to maintain the original order)
			return self.depthFirst(randItem,rootName,mainList,attribElemList)

		elif child == []:
			#print "\n"
			#print "It's a Leaf Node"
			lineage=[]
			attribElemList.append(node)
			#print "node(leaf)=",node
			mainList=self.printLineage(root,rootName,lineage,mainList,attribElemList,"firstAttrib")
			#print "mainList(received)",mainList
			#print "attribElemList(received)",attribElemList

			if root.tagName == rootName:
				return mainList
			
			node = root.parentNode									## get Parent Node
			root.parentNode.removeChild(root)						## remove current node
			#print "Parent node=",node
			return self.depthFirst(node,rootName,mainList,attribElemList)

########################################################################################################################################
########################################################################################################################################

	def createDataExcel(self,file):
		import instanceCreationDataEXL
		
		obj = instanceCreationDataEXL.instanceCreationDataEXL()
		retVal = obj.main(file)

########################################################################################################################################
########################################################################################################################################

	def createControlExcel(self,file):
		print "Processing: ",file
		print "####: ", os.path.basename(file)
		print "####: ",os.path.dirname(file)
		
		file_name = os.path.basename(file)
		baseDir = os.path.dirname(file)
		
		DOMTree = xml.dom.minidom.parse(file)
		
		''' Get the root node ('Device') '''
		for child in DOMTree.childNodes:
			if child.nodeType == xml.dom.Node.ELEMENT_NODE:
				device = child


		''' Get Comment Node '''
		for child in DOMTree.childNodes:
			if child.nodeType == xml.dom.Node.COMMENT_NODE:
				comment = child


		try:
			watermarkText = comment.data
		except UnboundLocalError:
			print "No watermark exists in this file"
			watermarkText = "No watermark"


		wb = Workbook()
		ws = wb.active
		ws.title=device.tagName
		#wb.remove_sheet(ws)				## remove the sheet "X2EMaster sheet in workBook"


		''' create excel sheets in a workbook '''
		for services in device.childNodes:
			if services.nodeType == xml.dom.Node.ELEMENT_NODE:
				ws = wb.create_sheet()
				ws.title = services.tagName


		''' ## get all the attribute names available in the XML '''
		attribNameList=getAttributesList.main(file)
		#print "attribNameList(from the XML)=",attribNameList
		#####################################################################################

		''' --- DEVICE :: First sheet of excel, root (or Device in this XML file) '''
		ws = wb.get_sheet_by_name(name = device.tagName)
		
		ws.cell(row=ATTRIB_NAME_ROW,column=1).value = 'PATH'
		ws.cell(row=ATTRIB_NAME_ROW,column=3).value = 'Source'
		ws.cell(row=ATTRIB_NAME_ROW+1,column=3).value = 'TR181/TR104'
		ws.cell(row=ATTRIB_NAME_ROW+2,column=3).value = 'LQ/Cust.'

		offsetRoot=0
		for attribNames in attribNameList:
			if attribNames == 'description':
				ws.cell(row=ATTRIB_NAME_ROW,column=2).value = 'Description'
			
			if attribNames != 'description':
				ws.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_NAME_COLUMN+offsetRoot).value = attribNames
				offsetRoot+=1
		
		XmlExlWatermark.insertInExcel(ws,PATH_ROW,PATH_COLUMN,watermarkText)

		ws.cell(row=PATH_ROW,column=PATH_COLUMN).value = device.tagName
		ws.cell(row=ATTRIB_NAME_ROW,column=4).value='defaultValue'
		ws.cell(row=PATH_ROW,column=4).value='--'
		ws.cell(row=ATTRIB_NAME_ROW,column=5).value='LQ_SUPP'
		ws.cell(row=PATH_ROW,column=5).value='Y'
		
		for colDumm in range(len(attribNameList)):
			ws.cell(row=PATH_ROW,column=colDumm+ATTRIB_NAME_COLUMN).value='--'

		attributes=device.attributes.items()
		
		for items in attributes:
			nameColIndex=0
			name=items[0]
			value=items[1]
			
			if name == 'description':
				ws.cell(row=PATH_ROW, column=2).value = value
			
			if name != 'description':
				while ws.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_NAME_COLUMN+nameColIndex).value != name:
					nameColIndex+=1
				
				ws.cell(row=PATH_ROW,column=ATTRIB_VALUE_COLUMN+nameColIndex).value=value
		###----------------------------------------------------------------------------------###
		########################################################################################
		
		''' Start populating data in each excel sheet of a workbook. '''
		for services in device.childNodes:
			if services.nodeType == xml.dom.Node.ELEMENT_NODE:
				ws = wb.get_sheet_by_name(name = services.tagName)
				
				ws.cell(row=ATTRIB_NAME_ROW,column=1).value = 'PATH'
				ws.cell(row=ATTRIB_NAME_ROW,column=3).value = 'Source'
				ws.cell(row=ATTRIB_NAME_ROW+1,column=3).value = 'TR181/TR104'
				ws.cell(row=ATTRIB_NAME_ROW+2,column=3).value = 'LQ/Cust.'
				
				offset=0							## write attribute Names
				for attribNames in attribNameList:
					if attribNames == 'description':
						ws.cell(row=ATTRIB_NAME_ROW,column=2).value = 'Description'
					
					if attribNames != 'description':
						ws.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_NAME_COLUMN+offset).value = attribNames
						offset+=1
				
				ws.cell(row=ATTRIB_NAME_ROW,column=4).value = 'defaultValue'
				ws.cell(row=ATTRIB_NAME_ROW,column=5).value='LQ_SUPP'
				
				root=services
				rootName=services.tagName
				mainList=[]
				attribElemList=[]
				listToWriteSheet = self.depthFirst(root,rootName,mainList,attribElemList)
				#print "listToWriteSheet=",listToWriteSheet,"\n"
				listToWriteSheet.reverse()								## two consecutive lists to be maintained in order
				attribElemList.reverse()								## attribElement List in order with listToWriteSheet
				index=0
				
				for i in listToWriteSheet:
					#print "i=",i
					#print "attribElemList=",attribElemList[index]		## element node instance
					
					ws.cell(row=PATH_ROW+index,column=PATH_COLUMN).value=i							## write the element lists
					ws.cell(row=PATH_ROW+index,column=PATH_COLUMN+2).value='TR181'
					
					''' write dummy values to all the attributes -- in excel sheet '''
					for colDumm in range(len(attribNameList)+1):
						ws.cell(row=PATH_ROW+index,column=colDumm+ATTRIB_NAME_COLUMN-1).value='--'
					
					
					''' write LANTIQ_SUPPORT = 1 in the respective column '''
					ws.cell(row=PATH_ROW+index,column=5).value = "Y"
					
					''' write the parameter Values '''
					try:
						elemNodeValue=attribElemList[index].firstChild.data
						if elemNodeValue != None:
							ws.cell(row=PATH_ROW+index,column=4).value=elemNodeValue
					except AttributeError:
						NodeValueDontExists=True
					
					''' write the attributes '''
					attrList = attribElemList[index].attributes.items()
					#print "attrList=",attrList
					
					for items in attrList:
						nameColIndex=0
						name=items[0]
						value=items[1]
						
						if name == 'description':
							ws.cell(row=PATH_ROW+index,column=2).value = value
						
						if name != 'description':
							while ws.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_NAME_COLUMN+nameColIndex).value != name:
								nameColIndex+=1
						
							#print "nameColIndex=",nameColIndex
							ws.cell(row=PATH_ROW+index,column=ATTRIB_VALUE_COLUMN+nameColIndex).value=value
					
					index+=1
		
		file = os.path.join(baseDir, file_name.split(".")[0]+".xlsx")
		print "file(toEXL##)=",file
		
		wb.save(file)
		print "Successfully created: ",file,"\n"

		''' Error handling '''
		returnVal = openpyxl_xlsxformatter(file)
		
		returnVal = 1
		if returnVal == 0:
			print "Status: Formatting successful"
			excelSplitVal = excelServiceSplitter(file)
		else:
			sys.exit(0)
		
		if excelSplitVal == 0:
			print "Status: Success"
			os.remove(file)