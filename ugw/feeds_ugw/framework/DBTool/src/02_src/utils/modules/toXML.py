
#********************************************************************************
#
#  Copyright (c) 2015
#  Lantiq Beteiligungs-GmbH & Co. KG
#  Lilienthalstrasse 15, 85579 Neubiberg, Germany 
#  For licensing information, see the file 'LICENSE' in the root folder of
#  this software module.
#
#********************************************************************************/

'''
This file creates an XML file from a Excel sheet.
Input to this file is an excel sheet
'''

## Import python modules
import xml.etree.ElementTree as ET
from lxml import etree
from xml.etree.ElementTree import ElementTree, dump,tostring

from openpyxl import load_workbook
import os,sys
from itertools import islice
from xml.etree.ElementTree import Element, SubElement, ElementTree, dump,tostring
from xml.etree import ElementTree

from xml.dom.minidom import *
import xml.dom.minidom

from removeAllAttributes import removeAllAttributes

sys.path.append('../common')
from common import *

sys.path.append('modules')
from excel_reference_values import *

sys.path.append(r'../action/')
from pathCorrect import pathCorrect

import addWaterMark

class toXML():

	def __init__(self):
		print "Initialising: toXML..."
	
	def reFormatXML(self,xmlFile):
		DOM = xml.dom.minidom.parse(xmlFile)
		f = open(xmlFile,"w")
		f.write(DOM.toprettyxml())
		f.close()
	
	def restructXML(self,xmlFile):
		''' Returns a restructured xml tree moving parameter nodes above child object nodes '''
		parser = etree.XMLParser(remove_blank_text=True) # lxml.etree only!
		tree = etree.parse(xmlFile, parser)
		root = tree.getroot()
		for child in root:
			for node in child.iter():
			#if no. of children is 0, that is a parameter node
				if len(node) ==  0:
					node.getparent().insert(0, node)
		
		newTree = tree
		new_root = newTree.getroot()
		completeName = xmlFile
		f = open(completeName,'w')
		f.write('<?xml version="1.0" ?>\n')
		f.write(tostring(new_root))
		f.close()
	
	def insertTagValue(self,xmlFile):
		#print "xmlFile(insertTagValue)",xmlFile
		f=open(xmlFile, 'rt')
		tree = ElementTree.parse(f)
		#print tree
		transRoot = tree.getroot()

		for node in tree.iter():
			#print "node=",node
			try:
				value=node.attrib['defaultValue']
				#print "value=", value
				
				if value != 'None':
					node.text = value
				
# TEMP BLOCK
#				if value == 'None':
#					value = ' '
#				
#				node.text=value
# TEMP BLOCK
			except KeyError:
				pass
		
		f = open(xmlFile,'w')
		f.write('<?xml version="1.0" ?>\n')
		f.write(tostring(transRoot))
		f.close()
		
	def removeParameterValue(self,xmlFile):
		#print "xmlFile(removeNodeValue)",xmlFile
		f=open(xmlFile, 'rt')
		tree = ElementTree.parse(f)
		#print tree
		transRoot = tree.getroot()
		
		for node in tree.iter():
			#print "node=",node
			try:
				##print node.attrib.pop('defaultValue')
				node.attrib.pop('defaultValue')
			except KeyError:
				pass
		
		f = open(xmlFile,'w')
		f.write('<?xml version="1.0" ?>\n')
		f.write(tostring(transRoot))
		f.close()
	
	def removeNoneAttribute(self,xmlFile):
		print "xmlFile(removeNoneAttribute)",xmlFile
		f=open(xmlFile, 'rt')
		tree = ElementTree.parse(f)
		#print tree
		transRoot = tree.getroot()

		for node in tree.iter():
			#print "node=",node
			attributesDict = node.attrib
			
			NoneElements=[]
			for key,value in attributesDict.iteritems():
				#print key,"--",value
				KEY = key
				VALUE = value
				if VALUE == 'None':
					NoneElements.append(KEY)					## Bug fixed -- removal of multiple attributes with values 'None'
			
			for keys in NoneElements:
				#print "keys=",keys
				attributesDict.pop(keys,'None')
			
			#print "\n"
		
		f = open(xmlFile,'w')
		f.write('<?xml version="1.0" ?>\n')
		f.write(tostring(transRoot))
		f.close()
		
	def addAttributes(self,node,dict):
		#print node
		#print dict
		for i in dict:
			if ((dict[i] != '--')):			## since 'parameterValue' is handled separately
				#print "name=",i,"val=",dict[i]
				value=dict[i]					## value
				name=i							## name
				node.setAttribute(name,str(value))					## BUG FIXED - if value is a number, it was throwing an error at setAttribute.
																	## AttributeError resolved by converting any contents to string format.
		#print "\n"

	## this function takes list as an input and build XML recursively
	def buildXML(self,list,index,dict='NULL',parent='NULL'):
		#print "list=",list
		#print "index(buildXML)=",index
		depth=len(list)
		#print "depth=",depth
		#print "dict=",dict

		##########################################################################################
		## This small loop deals with the service name and attributes like deviceInfo, Wifi, etc
		if depth == 1:
			element=self.services
			#print "element=",element
			#print "dict=",dict
			self.addAttributes(element,dict)

		### ENDS #################################################################################
		##########################################################################################

		## during the first iteration, the service name is taken
		if index==1:
			parent=self.services

		## if all the elements in the list are operated, it'll return back to main function
		if index==depth:
			return

		#print "elementUnderStudy=",list[index]
		#print "parent=",parent

		childList=[]											## contains child element instances
		childListName=[]										## contains child element names

		for child in parent.childNodes:
			if child.nodeType == xml.dom.Node.ELEMENT_NODE:
				#print "child=",child
				childList.append(child)
				childListName.append(child.tagName)

		#print "childList=",childList
		#print "childListName=",childListName

		## if element exists, it moves to next element in the list
		if list[index] in childListName:
			#print "element exists"
			indexParent = childListName.index(list[index])
			index+=1
			parent = childList[indexParent]
			return self.buildXML(list,index,dict,parent)

		## element doesn't exists and needs to be created, once created it moves to the next element in the list
		else:
			#print "create element"
			child=self.doc.createElement(list[index])
			#print "child=",child
			parent.appendChild(child)
			index+=1
			self.addAttributes(child,dict)
			parent=child
			return self.buildXML(list,index,dict,parent)

###########################################################################################################################
###########################################################################################################################
	def createDataXML(self,file):
		import instanceCreationDataXML
		
		obj = instanceCreationDataXML.instanceCreation()
		xmlFile,fileType,access = obj.main(file)

		self.insertTagValue(xmlFile)								## function inserts parameterValue
		self.removeParameterValue(xmlFile)							## function removes attribute 'parameterValue'
		self.removeNoneAttribute(xmlFile)
		
		import addWaterMark
		handler = addWaterMark.waterMark()
		handler.addWaterMark(xmlFile,fileType,access)
		
		print "Successfully created data XML: ",xmlFile
		
###########################################################################################################################
###########################################################################################################################
	
	def createControlXML(self,file):
		print "Received file(toXML)=",file
		print "####: ", os.path.basename(file)
		print "####: ",os.path.dirname(file)
		
		XML_NAME = os.path.basename(file)
		baseDir = os.path.dirname(file)
		
		wb=load_workbook(file)
		self.doc = Document()

		##########################################################################################
		############ handle the root of the XML file here (in this instance, the root is 'Device')
		rootSheet=wb.get_sheet_names()[0]
		#print "rootSheet=",rootSheet
		self.root=self.doc.createElement(rootSheet)							## root element to build a tree
		self.doc.appendChild(self.root)										## first child 'device' is added
		#print "device=",self.device
		
		wsR=wb.get_sheet_by_name(name = rootSheet)
		attribDict={}
		
		waterMarkText = (wsR['A4'].comment).text
		print "waterMarkText: ",waterMarkText
		index=0

		while wsR.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_NAME_COLUMN+index):	#(1,6)
			if wsR.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_NAME_COLUMN+index).value != None:
				attribDict[wsR.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_NAME_COLUMN+index).value] = wsR.cell(row=ATTRIB_VALUE_ROW,column=ATTRIB_VALUE_COLUMN+index).value
				index+=1						## column increment for every change in attribute
			else:
				break

		self.addAttributes(self.root,attribDict)
		############----------------------------------------------------------------###############
		###########################################################################################

		workSheets= wb.get_sheet_names()

		for sheets in workSheets[1:]:										## Ignoring the first sheet,since its a root node
			print "service tab: ",sheets
			ws=wb.get_sheet_by_name(name = sheets)
			self.services=self.doc.createElement(sheets)
			self.root.appendChild(self.services)

			## to read the child nodes of services and take to a list
			index=0

			childOfServiceList=[]
			while ws.cell(row=PATH_ROW+index,column=PATH_COLUMN):
				if ws.cell(row=PATH_ROW+index,column=PATH_COLUMN).value != None:
					#print ws.cell(row=PATH_ROW+index,column=PATH_COLUMN).value						## prints cell value
					childOfServiceList.append(ws.cell(row=PATH_ROW+index,column=PATH_COLUMN).value)
					index+=1
				else:
					break

			finalList=[]
			
			rowIndex=0

			for i in childOfServiceList:
				'''this part takes the attributes and keep them in dict format'''
				attribDict={}

				index=0
				while ws.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_NAME_COLUMN+rowIndex):
					if ws.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_NAME_COLUMN+index).value != None:
						attribDict[ws.cell(row=ATTRIB_NAME_ROW,column=ATTRIB_NAME_COLUMN+index).value] = ws.cell(row=ATTRIB_VALUE_ROW+rowIndex,column=ATTRIB_VALUE_COLUMN+index).value
						index+=1									## column increment for every change in attribute
					else:
						break

				'''these two lines specifically reads the 'defaultValue' column of excel sheet'''
				attribDict[ws.cell(row=PARAMETER_NAME_ROW,column=PARAMETER_NAME_COLUMN).value] = ws.cell(row=PARAMETER_VALUE_ROW+rowIndex,column=PARAMETER_NAME_COLUMN).value
				
				list=i.split('.')
				self.buildXML(list,index=1,dict=attribDict)
				rowIndex+=1							## ROW increment for every change in element node
				
				#print "\n"

		xmlFile = os.path.join(baseDir, XML_NAME.split(".xlsx")[0] + ".xml")
		print "xmlFile=",xmlFile
		
		f=open(xmlFile,'w')
		f.write(self.doc.toxml())
		f.close()

		self.insertTagValue(xmlFile)								## function inserts parameterValue
		self.removeParameterValue(xmlFile)							## function removes attribute parameter Value
		self.removeNoneAttribute(xmlFile)
		self.restructXML(xmlFile)									## Brings the parameter node to the top
		self.reFormatXML(xmlFile)									## prepares good format when opened in text mode
		
		
		if waterMarkText != 'No watermark':
			fileType = waterMarkText.split(',')[0]
			access = waterMarkText.split(',')[1]
		else:
			fileType = 'schema'
			access = 'ReadOnly'
		
		
		handler = addWaterMark.waterMark()
		handler.addWaterMark(xmlFile,fileType,access)
		
		print "Successfully created control XML: ",xmlFile